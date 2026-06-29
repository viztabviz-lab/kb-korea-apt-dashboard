# -*- coding: utf-8 -*-
"""
KB부동산 데이터허브 - 전국(17개 시도) 아파트 시세(통계 시계열) 수집기
- 인증키 불필요 (data-api.kbland.kr 공개 통계 API 사용)
- 출력: data/kb_korea.json (대시보드용 compact JSON)
         data/kb_korea_시계열.csv (Tableau용 long-format)
         data/KB_전국아파트시세.xlsx (요약 + 지표별 시트)
실행: python kb_korea.py
"""
import csv
import json
import sys
import datetime as dt
from pathlib import Path

import requests
import urllib3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
sys.stdout.reconfigure(encoding="utf-8")

BASE = "https://data-api.kbland.kr/bfmstat/weekMnthlyHuseTrnd/"
HEADERS = {"User-Agent": "Mozilla/5.0"}
HERE = Path(__file__).resolve().parent
OUT = HERE / "data"
OUT.mkdir(exist_ok=True)

# 전국 + 17개 시도. KB API는 최상위 호출(지역코드 미지정) 시 전국/광역집계/시도를
# 한 번에 내려준다. 아래 집합에 든 지역만 골라 담는다(광역집계는 제외).
SIDO_ORDER = ["전국", "서울", "경기", "인천", "부산", "대구", "광주", "대전", "울산",
              "세종", "강원", "충북", "충남", "전북", "전남", "경북", "경남", "제주"]
SIDO_SET = set(SIDO_ORDER)

# 시도명 → 파일 slug (대시보드가 시도 선택 시 lazy-load 하는 시군구 파일명)
SLUG = {
    "서울": "seoul", "경기": "gyeonggi", "인천": "incheon", "부산": "busan",
    "대구": "daegu", "광주": "gwangju", "대전": "daejeon", "울산": "ulsan",
    "세종": "sejong", "강원": "gangwon", "충북": "chungbuk", "충남": "chungnam",
    "전북": "jeonbuk", "전남": "jeonnam", "경북": "gyeongbuk", "경남": "gyeongnam",
    "제주": "jeju",
}


def fetch(endpoint, **params):
    """KB API 호출 -> (날짜리스트, [{지역코드,지역명,dataList}]) 반환"""
    r = requests.get(BASE + endpoint, params=params, headers=HEADERS,
                     verify=False, timeout=30)
    body = r.json()["dataBody"]
    if str(body["resultCode"]) != "11000":
        raise RuntimeError(f"{endpoint} 실패: {body}")
    d = body["data"]
    return d["날짜리스트"], d["데이터리스트"]


def parse_date(s):
    s = str(s)
    if len(s) == 8:
        return dt.date(int(s[:4]), int(s[4:6]), int(s[6:8]))
    return dt.date(int(s[:4]), int(s[4:6]), 1)   # 월간 YYYYMM


def collect(endpoint, 지표, 거래구분, regions, **params):
    """한 지표를 long-format 레코드 리스트로 변환.
    regions: 포함할 지역명 집합(None이면 전체)"""
    dates, rows = fetch(endpoint, **params)
    dates_n = len(dates)
    recs = []
    for row in rows:
        name = row["지역명"]
        if regions is not None and name not in regions:
            continue
        for i, v in enumerate(row["dataList"][:dates_n]):
            if v is None:
                continue
            recs.append({
                "날짜": parse_date(dates[i]).isoformat(),
                "지역명": name,
                "지표": 지표,
                "거래구분": 거래구분,
                "값": round(float(v), 4),
            })
    return recs


def gather():
    all_recs = []
    print("· 매매 가격지수(주간) ...")
    all_recs += collect("priceIndex", "가격지수", "매매", SIDO_SET,
                        월간주간구분코드="02", 매물종별구분="01", 매매전세코드="01")
    print("· 전세 가격지수(주간) ...")
    all_recs += collect("priceIndex", "가격지수", "전세", SIDO_SET,
                        월간주간구분코드="02", 매물종별구분="01", 매매전세코드="02")
    print("· 매매 평균가격(월간, 만원) ...")
    all_recs += collect("avgPrc", "평균가격", "매매", SIDO_SET,
                        매물종별구분="01", 매매전세코드="01")
    print("· 전세 평균가격(월간, 만원) ...")
    all_recs += collect("avgPrc", "평균가격", "전세", SIDO_SET,
                        매물종별구분="01", 매매전세코드="02")
    print("· 전세가율(월간, %) ...")
    all_recs += collect("dealCntstTnantRato", "전세가율", "-", SIDO_SET,
                        매물종별구분="01")
    return all_recs


def write_csv(recs):
    path = OUT / "kb_korea_시계열.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["날짜", "지역명", "지표", "거래구분", "값"])
        w.writeheader()
        w.writerows(sorted(recs, key=lambda r: (r["지표"], r["거래구분"], r["지역명"], r["날짜"])))
    return path


def latest_snapshot(recs):
    """(지표,거래구분,지역명) 별 최신값"""
    latest = {}
    for r in recs:
        k = (r["지표"], r["거래구분"], r["지역명"])
        if k not in latest or r["날짜"] > latest[k]["날짜"]:
            latest[k] = r
    return latest


def write_excel(recs):
    path = OUT / "KB_전국아파트시세.xlsx"
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1f4e79")
    hdr_font = Font(color="FFFFFF", bold=True)

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(1, c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    # 1) 요약(최신) 시트
    latest = latest_snapshot(recs)
    ws = wb.active
    ws.title = "요약(최신)"
    cols = ["지역", "매매 평균가(억)", "전세 평균가(억)", "매매 가격지수",
            "전세 가격지수", "전세가율(%)"]
    ws.append(cols)
    for sido in SIDO_ORDER:
        def g(지표, 거래):
            r = latest.get((지표, 거래, sido))
            return r["값"] if r else None
        mae_avg = g("평균가격", "매매")
        jeon_avg = g("평균가격", "전세")
        ws.append([
            sido,
            round(mae_avg / 10000, 2) if mae_avg else "",   # 만원 -> 억
            round(jeon_avg / 10000, 2) if jeon_avg else "",
            g("가격지수", "매매") or "",
            g("가격지수", "전세") or "",
            g("전세가율", "-") or "",
        ])
    style_header(ws, len(cols))
    ws.column_dimensions["A"].width = 12
    for c in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15

    # 2) 지표별 wide 시트 (행=날짜, 열=지역)
    def wide_sheet(title, 지표, 거래):
        sub = [r for r in recs if r["지표"] == 지표 and r["거래구분"] == 거래]
        if not sub:
            return
        dates = sorted({r["날짜"] for r in sub})
        regions = [g for g in SIDO_ORDER if any(r["지역명"] == g for r in sub)]
        idx = {(r["날짜"], r["지역명"]): r["값"] for r in sub}
        ws = wb.create_sheet(title[:31])
        ws.append(["날짜"] + regions)
        for d in dates:
            ws.append([d] + [idx.get((d, g), "") for g in regions])
        style_header(ws, len(regions) + 1)
        ws.column_dimensions["A"].width = 12

    wide_sheet("매매 가격지수", "가격지수", "매매")
    wide_sheet("전세 가격지수", "가격지수", "전세")
    wide_sheet("전세가율", "전세가율", "-")
    wide_sheet("매매 평균가격(만원)", "평균가격", "매매")
    wide_sheet("전세 평균가격(만원)", "평균가격", "전세")

    wb.save(path)
    return path


def build_series(recs, 지표, 거래):
    sub = [r for r in recs if r["지표"] == 지표 and r["거래구분"] == 거래]
    dates = sorted({r["날짜"] for r in sub})
    regions = [g for g in SIDO_ORDER if any(r["지역명"] == g for r in sub)]
    idx = {(r["날짜"], r["지역명"]): r["값"] for r in sub}
    out = {"dates": dates}
    for g in regions:
        out[g] = [idx.get((d, g)) for d in dates]
    return out


def write_json(recs):
    """대시보드용 compact JSON"""
    path = OUT / "kb_korea.json"
    latest = latest_snapshot(recs)
    summary = []
    for sido in SIDO_ORDER:
        def g(지표, 거래):
            r = latest.get((지표, 거래, sido))
            return r["값"] if r else None
        ma = g("평균가격", "매매")
        je = g("평균가격", "전세")
        summary.append({
            "region": sido,
            "매매평균억": round(ma / 10000, 2) if ma else None,
            "전세평균억": round(je / 10000, 2) if je else None,
            "매매지수": g("가격지수", "매매"),
            "전세지수": g("가격지수", "전세"),
            "전세가율": g("전세가율", "-"),
        })
    weekly_dates = sorted({r["날짜"] for r in recs if r["지표"] == "가격지수"})
    monthly_dates = sorted({r["날짜"] for r in recs if r["지표"] == "평균가격"})
    payload = {
        "source": "KB부동산 데이터허브 주택가격동향조사(아파트)",
        "updated": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "asof": {
            "weekly": weekly_dates[-1] if weekly_dates else None,
            "monthly": monthly_dates[-1] if monthly_dates else None,
        },
        "summary": summary,
        "series": {
            "매매지수": build_series(recs, "가격지수", "매매"),
            "전세지수": build_series(recs, "가격지수", "전세"),
            "전세가율": build_series(recs, "전세가율", "-"),
            "매매평균": build_series(recs, "평균가격", "매매"),
            "전세평균": build_series(recs, "평균가격", "전세"),
        },
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    return path


# ---------------------------------------------------------------------------
# 시군구 단위(드릴다운): 시도별 지역코드로 호출하면 하위 시군구가 내려온다.
# KB는 시군구 단위로 매매/전세 가격지수·전세가율을 제공하며 평균가격은 미제공.
# 시도별로 data/sg/kb/<slug>.json 파일을 만들어 대시보드가 필요할 때 로드한다.
# ---------------------------------------------------------------------------

def fetch_sido_codes():
    """{시도명: 지역코드} (전국·광역집계 제외)"""
    _, rows = fetch("priceIndex", 월간주간구분코드="02", 매물종별구분="01", 매매전세코드="01")
    return {r["지역명"]: r.get("지역코드") for r in rows
            if r["지역명"] in SIDO_SET and r["지역명"] != "전국"}


def collect_sigungu(code):
    """한 시도(code) 하위 시군구의 매매/전세 가격지수·전세가율 레코드"""
    recs = []
    recs += collect("priceIndex", "가격지수", "매매", None,
                    월간주간구분코드="02", 매물종별구분="01", 매매전세코드="01", 지역코드=code)
    recs += collect("priceIndex", "가격지수", "전세", None,
                    월간주간구분코드="02", 매물종별구분="01", 매매전세코드="02", 지역코드=code)
    recs += collect("dealCntstTnantRato", "전세가율", "-", None,
                    매물종별구분="01", 지역코드=code)
    return recs


def build_series_named(recs, 지표, 거래):
    """build_series 의 시군구판: 지역명을 API 반환 순서대로 유지"""
    sub = [r for r in recs if r["지표"] == 지표 and r["거래구분"] == 거래]
    dates = sorted({r["날짜"] for r in sub})
    regions = []
    for r in sub:
        if r["지역명"] not in regions:
            regions.append(r["지역명"])
    idx = {(r["날짜"], r["지역명"]): r["값"] for r in sub}
    out = {"dates": dates}
    for g in regions:
        out[g] = [idx.get((d, g)) for d in dates]
    return out


def write_sigungu():
    """시도별 시군구 파일 작성 후 [{name, slug, regions}] 반환"""
    base = OUT / "sg" / "kb"
    base.mkdir(parents=True, exist_ok=True)
    codes = fetch_sido_codes()
    info = []
    for name in SIDO_ORDER:
        if name == "전국":
            continue
        code = codes.get(name)
        if not code:
            continue
        recs = collect_sigungu(code)
        if not recs:
            continue
        weekly = sorted({r["날짜"] for r in recs if r["지표"] == "가격지수"})
        payload = {
            "asof": {"weekly": weekly[-1] if weekly else None},
            "series": {
                "매매지수": build_series_named(recs, "가격지수", "매매"),
                "전세지수": build_series_named(recs, "가격지수", "전세"),
                "전세가율": build_series_named(recs, "전세가율", "-"),
            },
        }
        with open(base / f"{SLUG[name]}.json", "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        n = len(payload["series"]["매매지수"]) - 1
        info.append({"name": name, "slug": SLUG[name], "regions": n})
        print(f"  · {name}: 시군구 {n}개")
    return info


def write_manifest(recs, sg_info, asof):
    """다운로더 페이지(download.html)용 파일 목록 manifest"""
    def sz(rel):
        p = OUT / rel
        return p.stat().st_size if p.exists() else 0

    updated = asof.get("weekly") or dt.date.today().isoformat()
    items = [
        {"source": "KB부동산", "scope": "전국", "name": "전국·시도 시계열",
         "file": "data/kb_korea_시계열.csv", "format": "CSV",
         "bytes": sz("kb_korea_시계열.csv"),
         "detail": f"전국+17개 시도 · {len(recs):,}행 · 매매/전세 지수·평균가·전세가율 (long-format)"},
        {"source": "KB부동산", "scope": "전국", "name": "전국·시도 대시보드 데이터",
         "file": "data/kb_korea.json", "format": "JSON",
         "bytes": sz("kb_korea.json"),
         "detail": "대시보드용 compact JSON (요약 + 시계열)"},
    ]
    for s in sg_info:
        items.append({
            "source": "KB부동산", "scope": "시군구", "name": f"{s['name']} 시군구",
            "file": f"data/sg/kb/{s['slug']}.json", "format": "JSON",
            "bytes": sz(f"sg/kb/{s['slug']}.json"),
            "detail": f"{s['name']} {s['regions']}개 시군구 · 매매/전세 지수·전세가율",
        })
    payload = {"source": "KB부동산", "updated": updated, "items": items}
    with open(OUT / "downloads_kb.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)


def main():
    print(f"[KB 전국 아파트 시세 수집] {dt.datetime.now():%Y-%m-%d %H:%M}")
    recs = gather()
    csv_path = write_csv(recs)
    json_path = write_json(recs)
    last = max(r["날짜"] for r in recs)
    print(f"\n완료 · 레코드 {len(recs):,}건 · 최신 기준일 {last}")
    print(f"  - {csv_path}")
    print(f"  - {json_path}")
    print("· 시군구 단위(드릴다운) 수집 ...")
    sg_info = write_sigungu()
    sg_total = sum(s["regions"] for s in sg_info)
    print(f"  - data/sg/kb/*.json (시군구 총 {sg_total}개)")
    write_manifest(recs, sg_info, {"weekly": last})
    print("  - data/downloads_kb.json")
    try:
        xlsx_path = write_excel(recs)
        print(f"  - {xlsx_path}")
    except PermissionError:
        print("  - (xlsx 건너뜀: 파일이 Excel에서 열려 있음)")


if __name__ == "__main__":
    main()
